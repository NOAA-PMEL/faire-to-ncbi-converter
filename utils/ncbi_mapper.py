import pandas as pd
import shutil
import frontmatter
import requests
import base64
import tempfile
import re
import numpy as np
import yaml
from utils.ncbi_mapping_data import faire_to_ncbi_units, ncbi_faire_to_ncbi_column_mappings_exact, ncbi_faire_sra_column_mappings_exact
from openpyxl import load_workbook
from pathlib import Path

pd.set_option('future.no_silent_downcasting', True)

class NCBIMapper:

    ncbi_sample_excel_template_path = 'MIMARKS.survey.water.6.0.xlsx'
    ncbi_sra_excel_template_path = 'SRA_metadata.xlsx'
    ncbi_organism = "marine metagenome"
    ncbi_sample_sheet_name = "MIMARKS.survey.water.6.0"
    ncbi_sra_sheet_name = 'SRA_data'
    ncbi_sample_header = 11
    ncbi_sra_header = 0
    ncbi_library_strategy = 'AMPLICON'
    ncbi_library_source = 'METAGENOMIC'
    ncbi_library_selection = 'PCR'
    ncbi_library_layout = 'Paired-end'
    ncbi_file_type = 'fastq'
    faire_sample_name_col = "samp_name"
    faire_sample_compose_of_col = "sample_composed_of"
    ncbi_bioprojec_col_name = 'bioproject_accession'
    faire_missing_values = ["not applicable: control sample",
                            "not applicable: sample group",
                            "not applicable",
                            "missing: not collected: synthetic construct",
                            "missing: not collected: lab stock",
                            "missing: not collected: third party data",
                            "missing: not collected",
                            "missing: not provided",
                            "missing: restricted access: endangered species", 
                            "missing: restricted access: human-identifiable", 
                            "missing: restricted access"
                            ]
    faire_samp_name_col = "samp_name"
    ncbi_samp_name_col = "*sample_name"

    def __init__(self, config_file: str):
        
        self.config_file = self.load_config(config_file)
        self.ncbi_sample_excel_save_path = Path(self.config_file.get('ncbi_sample_excel_save_path'))
        self.ncbi_sra_excel_save_path = Path(self.config_file.get('ncbi_sra_excel_save_path'))
        self.assay_type = self.config_file.get('assay_type')
        self.faire_sample_df, self.faire_experiment_run_df = self.prepare_dfs()
        self.ncbi_sample_template_df = self.load_ncbi_template_as_df(file_path=self.ncbi_sample_excel_template_path, sheet_name=self.ncbi_sample_sheet_name, header=self.ncbi_sample_header)
        self.library_prep_bebop = self.retrive_github_bebop(owner=self.config_file['library_prep_info'].get('owner'), repo=self.config_file['library_prep_info'].get('repo'), file_path=self.config_file['library_prep_info'].get('file_path'))
        self.ncbi_bioproject_dict = self.create_ncbi_accession_dict_project_and_biosample(id_prefix='PRJNA')
        self.ncbi_biosample_dict = self.create_ncbi_accession_dict_project_and_biosample(id_prefix='SAMN')
        self.ncbi_srr_dict = self.create_ncbi_accession_dict_project_and_biosample(id_prefix='SRR')
        self.geo_loc_dict = self.create_geo_loc_dict()
       
    def load_config(self, config_path):
        # Load configuration yaml file
    
        with open(config_path, 'r') as f:
            return yaml.safe_load(f)
        
    def create_ncbi_submission(self):
        # Will output two excel files

        # Create NCBI Sample Template and fill out and save
        final_ncbi_sample_df = self.get_ncbi_sample_df()
        self.save_to_excel_template(template_path=self.ncbi_sample_excel_template_path,
                                                    ncbi_excel_save_path=self.ncbi_sample_excel_save_path, 
                                                    sheet_name=self.ncbi_sample_sheet_name,
                                                    header=self.ncbi_sample_header,
                                                    final_ncbi_df=final_ncbi_sample_df)
        
        # Creat NCBI SRA Template and fill out and save
        final_sra_df = self.get_sra_df()
        self.save_to_excel_template(template_path=self.ncbi_sra_excel_template_path,
                                    ncbi_excel_save_path=self.ncbi_sra_excel_save_path,
                                    sheet_name=self.ncbi_sra_sheet_name,
                                    header=self.ncbi_sra_header,
                                    final_ncbi_df=final_sra_df)
    
    def create_osu_ncbi_submission(self):
        
        exp_run_groups, samp_groups = self.split_osu_dfs_by_submission_type()

        for submission_type, sample_df in samp_groups.items():
            # overwrite faire_experiment_run_df and faire_sample_df
            self.faire_experiment_run_df = exp_run_groups.get(submission_type)
            self.faire_sample_df = sample_df


            # Create new excel file save path: 
            new_dir = self.ncbi_sample_excel_save_path.parent / submission_type
            new_sample_file_name = f"{submission_type}_{self.ncbi_sample_excel_save_path.name}"
            new_sample_path = new_dir / new_sample_file_name
            # create directory if it doesn't already exist
            new_dir.mkdir(parents=True, exist_ok=True)

            # process new dfs for ncbi submission
            # First sample df
            final_ncbi_sample_df = self.get_ncbi_sample_df()
            self.save_to_excel_template(template_path=self.ncbi_sample_excel_template_path,
                                                    ncbi_excel_save_path=new_sample_path, 
                                                    sheet_name=self.ncbi_sample_sheet_name,
                                                    header=self.ncbi_sample_header,
                                                    final_ncbi_df=final_ncbi_sample_df)

        for submission_type, exp_run_df in exp_run_groups.items():
            
            new_dir = self.ncbi_sample_excel_save_path.parent / submission_type
            new_expRun_file_name = f"{submission_type}_{self.ncbi_sra_excel_save_path.name}"
            new_expRun_path = new_dir/ new_expRun_file_name
            self.faire_experiment_run_df = exp_run_df
            
            final_sra_df = self.get_sra_df()
            self.save_to_excel_template(template_path=self.ncbi_sra_excel_template_path,
                                    ncbi_excel_save_path=new_expRun_path,
                                    sheet_name=self.ncbi_sra_sheet_name,
                                    header=self.ncbi_sra_header,
                                    final_ncbi_df=final_sra_df)

    def split_osu_dfs_by_submission_type(self):
        # splits the self.samp_df and self.exp_run_df into separate data frames by submission type (e.g. single_direct, nan, etc.)
        # and returns to dictionary, the samp_groups and the exp_run_groups
        # replace NaN values for submission_type with 'NEW'
        
        mapped_biosamp_accessions = self.faire_experiment_run_df[self.faire_samp_name_col].map(self.ncbi_biosample_dict)
        has_biosamp_accession_mask = mapped_biosamp_accessions.notna()
        self.faire_experiment_run_df['group_key'] = np.where(
            has_biosamp_accession_mask,
            'has_biosamp_accession',
            'no_biosamp_accession'
        )
        
        # Step 1: split exp_run_df by submission type (inlcuding NaN or empty)
        exp_run_groups = {}
        for ncbi_biosample_status, group in self.faire_experiment_run_df.groupby('group_key', dropna=False):
            exp_run_groups[ncbi_biosample_status] = group

        # Step 2: Get the unique group/key ncbi_submission statuses to know how to split sample_df
        ncbi_biosample_statuses = self.faire_experiment_run_df['group_key'].unique() 

        # Step 3: split sample_df based on which sample names correspond to each submission value
        samp_groups = {}
        for ncbi_biosample_status in ncbi_biosample_statuses:
            # Get samp names that correspond to this submission value
            samp_names_for_submission = self.faire_experiment_run_df[self.faire_experiment_run_df['group_key'] == ncbi_biosample_status][self.faire_samp_name_col].unique()

            # filter samp_df to only include rows with these samp_names
            samp_groups[ncbi_biosample_status] = self.faire_sample_df[self.faire_sample_df[self.faire_samp_name_col].isin(samp_names_for_submission)]

        return exp_run_groups, samp_groups
    
    def load_ncbi_template_as_df(self, file_path: str, sheet_name: str, header: int) -> pd.DataFrame:
        # Load FAIRe excel template as a data frame based on the specified template sheet name
        
        return pd.read_excel(file_path, sheet_name=sheet_name, header=header)
    
    def concat_csvs(self, list_of_csv_paths: list) -> pd.DataFrame:
        """
        Reads a list of csvs and concats them into one data frame.
        """
        list_of_dfs = []
        for csv in list_of_csv_paths:
            df = pd.read_csv(csv)
            list_of_dfs.append(df)

        return pd.concat(list_of_dfs, ignore_index=True)
        
    def prepare_dfs(self) -> pd.DataFrame:
        # removes all FAIRe missing values and returns empty values
        # and concatenates the various sample data frames, and drops any samples from sample_df that aren't part of run to submit.
        sample_df_concated = self.concat_csvs(list_of_csv_paths=self.config_file.get('sample_metadata_project_csvs'))
        exp_df_concated = self.concat_csvs(list_of_csv_paths=self.config_file.get('run_metadata_csvs'))

        # Filter out samples from sample_df that are not part of this run/ncbi submission
        samp_df_filtered = self._filter_samp_df_by_samp_name(samp_df=sample_df_concated, associated_seq_df=exp_df_concated)
        
        # Check that all samples in the run df are present in the sample df - will throw error if not.
        self.check_all_samps_are_present_in_exp_df(final_sample_df=samp_df_filtered, exp_run_df=exp_df_concated)

        # Clean both sample_df and exp_df to making missing values just empty strings
        samp_df_clean = samp_df_filtered.fillna('').replace(self.faire_missing_values, '')
        run_df_clean = exp_df_concated.fillna('').replace(self.faire_missing_values, '')
        
        return samp_df_clean, run_df_clean
    
    def _filter_samp_df_by_samp_name(self, samp_df: pd.DataFrame, associated_seq_df: pd.DataFrame) -> pd.DataFrame:
        # filter sample dataframe to keep only rows where sample names exist in associated seq data frame
        
        missing_samples = []
        # Get unique sample name values from associated Dataframe
        exp_valid_samps = set(associated_seq_df[self.faire_sample_name_col].unique())

        # Also valid samples are pooled subsamples that won't necessarily exist in the experiment run metadata
        pooled_samp_dict = self.get_dict_of_pooled_samps(sample_df=samp_df)
        other_valid_samps = set()
        for pooled_sample, samps_that_were_pooled in pooled_samp_dict.items():
            for samp in samps_that_were_pooled:
                if samp.strip() in set(samp_df[self.faire_sample_name_col]):
                    other_valid_samps.add(samp)

        # combine valid samps
        valid_samp_names = exp_valid_samps | other_valid_samps

        # Identify missing sample names (those in primary but not in associated seq runs)
        cruise_sample_samp_names = samp_df['samp_name'].unique()
        missing_samples = [name for name in cruise_sample_samp_names if name not in valid_samp_names]

        # Save missing samples to logging dir
        if len(missing_samples) > 1:
            self.save_samples_dropped_to_logging(dropped_samples=missing_samples)
        
        filtered_df = samp_df[samp_df['samp_name'].isin(valid_samp_names)].copy()

        return filtered_df
    
    def check_all_samps_are_present_in_exp_df(self, final_sample_df: pd.DataFrame, exp_run_df: pd.DataFrame) -> pd.DataFrame:
        # Checks that all samples in the exp_df exist in the samp_df, if not, will throw an error!

        # Get list of reference names
        reference_names = set(final_sample_df[self.faire_sample_name_col].unique())

        # Create mask of rows to keep
        mask = exp_run_df[self.faire_sample_name_col].isin(reference_names)

        # Get dropped samples
        dropped_samples = exp_run_df.loc[~mask, self.faire_sample_name_col].unique()

        if len(dropped_samples) != 0:
            raise ValueError(f"\033[31mThere are samples missing from the sample csvs that are in this run! Please add sample metadata for {dropped_samples}!\033[0m")
    
    def get_dict_of_pooled_samps(self, sample_df: pd.DataFrame) -> dict:
        """
        Gets a dictionary of the pooled sample name: list of poooled samp names
        """
        # Get all rows where there is a list in the sample_compose_of col (pooled samples)
        mask = sample_df[self.faire_sample_compose_of_col].str.contains('|', regex=False)
        col_to_split = sample_df.loc[mask, self.faire_sample_compose_of_col]
        samples_pooled = col_to_split.str.split(' | ')
        pooled_samps = sample_df.loc[mask, self.faire_samp_name_col]
        pooled_dict = dict(zip(pooled_samps, samples_pooled))
        return pooled_dict

    def save_samples_dropped_to_logging(self, dropped_samples: list):
        """
        Takes a list of the dropped samples from the sample df (because they
        weren't part of the run - theoretically, unless there is a mismatch error
        with the sample names) and adds them to a csv in the associated logging
        folder called samples_not_part_of_run.csv
        """
        # Gets the directory where we are saving the data to (in the config file)
        main_save_dir = self.ncbi_sample_excel_save_path.parent 
        logging_dir = main_save_dir / 'logging'

        # Check if looging directory exists and create it if it doesn't.
        # parents = True ensures any missing parent directories are also created
        # exist_ok = True prevents an error if the directory already exists
        logging_dir.mkdir(parents=True, exist_ok=True)
        new_file_name = "dropped_samples_not_part_of_run.csv"
        logging_results_path = logging_dir / new_file_name

        # Create df and save 
        df = pd.DataFrame(dropped_samples)
        df.to_csv(logging_results_path, index=False, header=False)

    def get_ncbi_sample_df(self):

        # updated_df = pd.DataFrame()
        data_for_df = {}

        # First handle unit transormations
        for ncbi_col_name, faire_mapping in faire_to_ncbi_units.items():
            # Get FAIRe column mame
            faire_col = faire_mapping['faire_col']

            if faire_col not in self.faire_sample_df:
                continue
            else:
                # Check if we have a constant unit or a unit column
                if 'constant_unit_val' in faire_mapping:
                    unit_val = faire_mapping['constant_unit_val']
                    new_col_data = (self.faire_sample_df[faire_col].astype(str) + ' ' + unit_val).where(self.faire_sample_df[faire_col].astype(str).str.strip() != '', '')
                elif 'faire_unit_col' in faire_mapping:
                    unit_col = faire_mapping['faire_unit_col']
                    new_col_data = self.faire_sample_df[faire_col].astype(str) + ' ' + self.faire_sample_df[unit_col].astype(str).where(self.faire_sample_df[faire_col].astype(str).str.strip() != '', '')
                else:
                    new_col_data = self.faire_sample_df[faire_col]
                # Store the resulting sereis in the dictionary
                data_for_df[ncbi_col_name] = new_col_data

        # Second handle direct column mappings
        for old_col_name, new_col_name in ncbi_faire_to_ncbi_column_mappings_exact.items():
            if old_col_name in self.faire_sample_df:
                data_for_df[new_col_name] = self.faire_sample_df[old_col_name]

        # Third handle logic cases (depth and lat/lon)
        data_for_df['*depth'] = self.faire_sample_df.apply(
                lambda row: self.get_ncbi_depth(metadata_row=row),
                axis=1
            )
        data_for_df['*lat_lon'] = self.faire_sample_df.apply(
            lambda row: self.get_ncbi_lat_lon(metadata_row=row),
            axis=1
        )

        # Fourth add organism
        data_for_df['*organism'] = self.ncbi_organism

        # Fifth add description for PCR technical replicates and additional column - needed to differentiate their metadata for submission to be accepted
        data_for_df['description'] = self.faire_sample_df['samp_name'].apply(self.add_description_for_pcr_reps)
        data_for_df['technical_rep_id'] = self.faire_sample_df['samp_name'].apply(self.add_pcr_technical_rep)

        # Get Bioproject accession from exp run df and biosample acession (important for OSU runs). If value is left blank, add
        # new bioaccession number from the config.yaml file
        data_for_df[self.ncbi_bioprojec_col_name] = data_for_df[self.ncbi_samp_name_col].map(self.ncbi_bioproject_dict)
        data_for_df[self.ncbi_bioprojec_col_name] = data_for_df[self.ncbi_bioprojec_col_name].fillna(self.config_file.get('bioproject_accession'))

        # Add the sample_title
        data_for_df['sample_title'] = self.faire_sample_df.apply(lambda row: self.create_sample_title(metadata_row=row), axis=1)

        # --- Create the initial DataFrame from the collected data ---
        updated_df = pd.DataFrame(data_for_df)
     
        # drop technical_rep_id if its empty
        if 'technical_rep_id' in updated_df.columns and updated_df['technical_rep_id'].isnull().all():
            updated_df = updated_df.drop(columns=['technical_rep_id'])

        # 6th add additional column in FAIRe sample df that do not exist in ncbi template
        final_ncbi_df = self.add_additional_faire_cols(faire_samp_df=self.faire_sample_df, updated_ncbi_df=updated_df)

        # Add biosample_accession - did not like it when I added it before this.
        final_ncbi_df['biosample_accession'] = final_ncbi_df[self.ncbi_samp_name_col].map(self.ncbi_biosample_dict)

        # Drop empty columns
        last_final_ncbi_df = self.drop_empty_cols(final_ncbi_df)
        
        return last_final_ncbi_df

    def get_sra_df(self) -> pd.DataFrame:

        updated_df = pd.DataFrame()

        # First handle direct mappings
        for old_col_name, new_col_name in ncbi_faire_sra_column_mappings_exact.items():
            if old_col_name in self.faire_experiment_run_df:
                updated_df[new_col_name] = self.faire_experiment_run_df[old_col_name]

        # Second add values that are the same across all samples
        updated_df['library_strategy'] = self.ncbi_library_strategy
        updated_df['library_source'] = self.ncbi_library_source
        updated_df['library_selection'] = self.ncbi_library_selection
        updated_df['library_layout'] = self.ncbi_library_layout
        updated_df['platform'] = self.library_prep_bebop['platform']
        updated_df['instrument_model'] = self.library_prep_bebop['instrument']
        updated_df['design_description'] = f"Sequencing performed at {self.library_prep_bebop['sequencing_location']}"
        updated_df['filetype'] = self.ncbi_file_type
        updated_df['title'] = updated_df.apply(
            lambda row: self.create_SRA_title(metadata_row=row),
            axis=1
        )

        # # Add biosample_accession - did not like it when I added it before this.
        updated_df['biosample_accession'] =  updated_df['sample_name'].map(self.ncbi_biosample_dict)
        # drop biosample_accession if its empty (for non-osu runs)
        if 'biosample_accession' in updated_df.columns and updated_df['biosample_accession'].isnull().all():
            updated_df = updated_df.drop(columns=['biosample_accession'])
       
        # Add srr_accession - did not like it when I added it before this.
        updated_df['srr_accession'] = updated_df['library_ID'].map(self.ncbi_srr_dict)
        # drop srr_accession if its empty (for non-osu runs)
        if 'srr_accession' in updated_df.columns and updated_df['srr_accession'].isnull().all():
            updated_df = updated_df.drop(columns=['srr_accession'])

        return updated_df

    def load_beBop_yaml_terms(self, path_to_bebop: str):
        # read BeBOP yaml terms
        with open(path_to_bebop, 'r', encoding='utf-8') as f:
            post = frontmatter.load(f)
            return post
        
    def retrive_github_bebop(self, owner: str, repo: str, file_path: str):
        url = f"https://api.github.com/repos/{owner}/{repo}/contents/{file_path}"

        try:
            response = requests.get(url)
            response.raise_for_status()
            data = response.json()

            if 'content' in data:
                # Decode base64 to get the raw markdown file
                base64_content = data['content'].replace('\n', '').replace(' ', '')
                markdown_content = base64.b64decode(base64_content).decode('utf-8')
                
                with tempfile.NamedTemporaryFile(mode='w', suffix='.md', delete=True, encoding='utf-8') as temp_file:
                    temp_file.write(markdown_content)
                    temp_file_path = temp_file.name
                    post = self.load_beBop_yaml_terms(path_to_bebop=temp_file_path)
                    return post.metadata
        
        except requests.exceptions.RequestException as e:
            print(f"Error fetching bebop: {e}")
            return None
        
    def get_faire_cols_that_map_to_ncbi(self) -> list:
        # Creates a list of all the columns in FAIRe that map to NCBI (from the .lists)
        faire_maps_to_ncbi_cols = []

        # Add unit/and constant value faire cols
        for key, details in faire_to_ncbi_units.items():
            if "faire_col" in details:
                faire_maps_to_ncbi_cols.append(details["faire_col"])
            if "faire_unit_col" in details:
                faire_maps_to_ncbi_cols.append(details["faire_unit_col"])

        # Add exact mapping faire cols
        for key in ncbi_faire_to_ncbi_column_mappings_exact.keys():
            faire_maps_to_ncbi_cols.append(key)

        return faire_maps_to_ncbi_cols
    
    def get_additional_faire_unit_cols_dict(self, faire_cols: list) -> dict:
        # Creates a dictionary of the additional faire_cols with that have corresponding unit cols and creates a dict with faire col as key and unit col as value
        unit_col_dict = {}
        for col in faire_cols:
            unit_col = f"{col}_unit"
            units_col = f"{col}_units"
            if unit_col in faire_cols:
                unit_col_dict[col] = unit_col
            if units_col in faire_cols:
                unit_col_dict[col] = units_col

        # remove columns from list that exist in dictionary
        to_remove = set(unit_col_dict.keys()) 
        to_remove.update(unit_col_dict.values())
        filtered_faire_list = [col for col in faire_cols if col not in to_remove]
        
        return unit_col_dict, filtered_faire_list

    def add_additional_faire_cols(self, faire_samp_df: pd.DataFrame, updated_ncbi_df: pd.DataFrame) -> pd.DataFrame:
        # Adds additional FAIRe columns not in the NCBI sample template
       
        # Get unit cols and regular cols - additional_reg_faire_cols is a list of regular columsn (not unit corresponding ones)
        # And additional_cols_units_dict is columns with unit columsn
        faire_cols_that_map = set(self.get_faire_cols_that_map_to_ncbi())
        faire_cols = set(faire_samp_df.columns)
        additional_faire_cols = list(faire_cols - faire_cols_that_map)
        cols = self.get_additional_faire_unit_cols_dict(faire_cols=additional_faire_cols)
        additional_cols_units_dict = cols[0]
        additional_reg_faire_cols = cols[1]

        new_col_data = {}
        # Add unit cols first
        for key, value in additional_cols_units_dict.items():
            new_col_data[key] = faire_samp_df[key].astype(str) + ' ' + faire_samp_df[value].astype(str)

        # Then add regular columns (those with no corresponding units)
        for faire_col in additional_reg_faire_cols:
            new_col_data[faire_col] = faire_samp_df[faire_col]

        # Create the data frame of new columns and concat once
        if new_col_data:
            new_cols_df = pd.DataFrame(new_col_data, index=updated_ncbi_df.index)
            # Concat the new columns to the existing DatFrame in one go (axis=1 fo columns)
            final_ncbi_df = pd.concat([updated_ncbi_df, new_cols_df], axis=1)
        else:
            final_ncbi_df = updated_ncbi_df

        # Drop columns that are empty for all rows that were additional faire columns
        columns_to_drop = []
        for col in additional_faire_cols:
            if col in final_ncbi_df.columns:
                if final_ncbi_df[col].isna().all() or final_ncbi_df[col].astype(str).str.strip().eq('').all():
                    columns_to_drop.append(col)

        if columns_to_drop:
            final_ncbi_df = final_ncbi_df.drop(columns=columns_to_drop)

        return final_ncbi_df

    def get_ncbi_depth(self, metadata_row: pd.Series) -> str:
        # Get the ncbi formatted value for depth, which is the interval of minimumDepthInMeters - maximumDepthInMeters
        min_depth = metadata_row['minimumDepthInMeters']
        max_depth = metadata_row['maximumDepthInMeters']

        # If min_depth and max_depth are not the same, report as interval
        if min_depth != max_depth and max_depth != '':
            ncbi_depth = min_depth + ' ' + 'm' + ' - ' + max_depth + ' ' + 'm'
        # if min_depth and max_depth are the same, report just one because they will be the same
        elif min_depth == max_depth and max_depth != '':
            ncbi_depth = min_depth + ' ' + 'm'
        else:
            if min_depth == '' and max_depth == '': # for controll samples
                return ''
            else:
                raise ValueError(f"Something wrong with the depth. Min depth is {min_depth} and max_depth is {max_depth}. {metadata_row}")

        return ncbi_depth

    def get_ncbi_lat_lon(self, metadata_row: pd.Series) -> str:
        try:
            lat = float(metadata_row['decimalLatitude'])
            lon = float(metadata_row['decimalLongitude'])

            lat_dir = 'N' if lat >= 0 else 'S'
            lon_dir = 'E' if lon >= 0 else 'W'

            # Take absolute values
            lat_abs = abs(lat)
            lon_abs = abs (lon)

            lat_formatted = f"{lat_abs:.4f} {lat_dir}"
            lon_formatted = f"{lon_abs:.4f} {lon_dir}"

            ncbi_lat_lon = f"{lat_formatted} {lon_formatted}"

            return ncbi_lat_lon
        except:
            if metadata_row['decimalLatitude'] == '' and metadata_row['decimalLongitude'] == '':
                return ''
            else:
                raise ValueError(f"Something wrong with lat lon and can't transform for sample: {metadata_row[self.faire_samp_name_col]} with lat/lon {metadata_row['decimalLatitude']}/{metadata_row['decimalLongitude']}")
    
    def add_description_for_pcr_reps(self, sample_name: str) -> str:
        # Add a description for the PCR technical replicates (they have to be distinguishable to submit)
        if 'PCR' in sample_name:
            samp_parts = sample_name.split('.')
            # PCR number is last character of the last part
            pcr_num = samp_parts[-1][-1]
            original_sample = '.'.join(samp_parts[:-1])
            return f"PCR technical replicate number {pcr_num} of {original_sample}"
        else:
            return ''

    def add_pcr_technical_rep(self, sample_name: str) -> str:
        if 'PCR' in sample_name:
            samp_parts = sample_name.split('.')
            # PCR number is last character of the last part
            pcr_num = samp_parts[-1][-1]
            return pcr_num
        else:
            return ''
    
    def save_to_excel_template(self, template_path: str, ncbi_excel_save_path: str, sheet_name: str, header: int, final_ncbi_df: pd.DataFrame):

        # Copy template file at saved locationsample_ncbi_excel_save_path
        shutil.copy2(template_path, ncbi_excel_save_path)

        # Load the template to get the column structure
        template_df = pd.read_excel(template_path, sheet_name=sheet_name, header=header)
        template_columns = template_df.columns.tolist()
        
        # Load workbook to preserve formatting
        book = load_workbook(ncbi_excel_save_path)
        ws = book[sheet_name]

        # find column in final_ncbi_df that don't exist in template df (e.g. technical replicate)
        missing_cols = [col for col in final_ncbi_df.columns if col not in template_columns]

        # Add missing columns
        if missing_cols:
            next_col_idx = len(template_columns) + 1
            for i, col_name in enumerate(missing_cols):
                # Add header for new column
                ws.cell(row=header+1, column=next_col_idx + i, value=col_name)
                template_columns.append(col_name)

        for col_idx, template_col_name in enumerate(template_columns, start=1):
            if template_col_name in final_ncbi_df.columns:
                #write data for this column
                column_data = final_ncbi_df[template_col_name]
                for row_idx, value in enumerate(column_data, start = header+2):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
        # save workbook
        book.save(ncbi_excel_save_path)

        # # Save as tsv
        # df = pd.read_excel(sample_ncbi_excel_save_path, sheet_name=sheet_name)
        # df.to_csv(sample_ncbi_excel_save_path.replace('.xlsx', '.tsv'), sep='\t', index=False)

        print(f'NCBI sample saved to {ncbi_excel_save_path}')

    def drop_empty_cols(self, sample_df: pd.DataFrame) -> pd.DataFrame:
        # Drop empty columns. Regex pattern matches any cell containing zero or more whitespace characters (\s*) between the start (^) and end ($) of the string. This ensures truly empty or blank-space-only cells are replaced.
        df_cleaned = sample_df.replace(r'^\s*$', np.nan, regex=True)
        return df_cleaned

    def get_ncbi_bioproject_if_exists(self, metadata_row: pd.Series, id_prefix: str) -> str:
        # Uses the associatedSequences column in the FAIRe df to get the NCBI accession number 
        # # (e.g. for bioproject, srr, and biosample):
        try:
            associated_sequences = metadata_row['associatedSequences']
            associated_seqs = associated_sequences.split(' | ')
            for accession_num in associated_seqs:
                if id_prefix in accession_num:
                    match = re.search(rf'{id_prefix}\w+', accession_num)
                    if match:
                        ncbi_accession_id = match.group()
                        return ncbi_accession_id
        except:
            pass

    def create_ncbi_accession_dict_project_and_biosample(self, id_prefix: str) -> dict:
        # Creates a dictionary of accessions for ncbi (bioproject, biosample) 
        # Don't use for SRA because some samples will have duplicate or more samp_names and the values will get overwritten. This is
        # only for PRJNA and SAMN
        if id_prefix == 'SRR':
           group_col = 'lib_id'
        else:
            group_col = self.faire_samp_name_col

        exp_df = self.faire_experiment_run_df.copy()
        exp_df[id_prefix] = exp_df.apply(
            lambda row: self.get_ncbi_bioproject_if_exists(metadata_row=row, id_prefix=id_prefix),
            axis=1 
        )

        # Handle duplicate samp_name values and prioritize non-Non accessions
        grouped = exp_df.groupby(group_col)[id_prefix]

        # Use a dictionary comprehension to find the first non-None value for each group
        if group_col == self.faire_samp_name_col:
            ncbi_accession_dict = {
                samp_name: group.dropna().iloc[0] if not group.dropna().empty else None
                for samp_name, group in grouped
            }
        elif group_col == 'lib_id':
            ncbi_accession_dict = dict(zip(exp_df[group_col], exp_df[id_prefix]))
        
        return ncbi_accession_dict

    def create_geo_loc_dict(self) -> dict:
        # create a dictionary of sample name as key and geo location portion as value
        # to be used eventually in the Title of the SRA samples

        geo_loc_samp_dict = {
            row[self.faire_samp_name_col]: row['geo_loc_name'].split(': ', 1)[1] if ': ' in row['geo_loc_name'] else row['geo_loc_name'] for _, row in self.faire_sample_df.iterrows()
        }

        return geo_loc_samp_dict

    def create_SRA_title(self, metadata_row: pd.Series) -> str:
        # Create the title for the SRA template. E.g. 16S metabarcoding of sewater samples from the Bering Sea
        samp_name = metadata_row['sample_name'] # matches what is in the SRA template column
        library_id = metadata_row['library_ID'] # matches what is in the SRA template column
        location = self.geo_loc_dict.get(samp_name)

        run = library_id.split('_')[-1]
        if 'osu' in run.lower():
            run = f"Run {run.upper()}"
        else:
            run = run.capitalize()
        
        if samp_name.startswith('E'): # may need to add additional functionality here
            if 'PPS' in samp_name:
                type_of_samp = 'PPS-collected seawater'
            else:
                type_of_samp = 'CTD-collected seawater'

            return f"Environmental DNA (eDNA) {self.assay_type} of {type_of_samp} in the {location}: {run}"

    def create_sample_title(self, metadata_row: pd.Series) -> str:
        """
        Create the sample_title for the sample rows, using the faire_df metdata rows
        """
        samp_name = metadata_row[self.faire_samp_name_col] # matches faire field
        location = metadata_row['geo_loc_name'].replace('USA: ', '') # maches faire_field
        assay_name = metadata_row['assay_name'] # matches faire field
        samp_category = metadata_row['samp_category']
        if samp_category == 'sample': # may need to add additional functionality here
            if 'PPS' in samp_name:
                type_of_samp = 'PPS-collected seawater'
            else:
                type_of_samp = 'CTD-collected seawater'
     
            sample_title = f"Environmental DNA (eDNA) {self.assay_type} of {type_of_samp} in the {location}: Assay(s) {assay_name} survey of sample {samp_name}"

        elif samp_category == 'negative control':
            neg_cont_type = metadata_row['neg_cont_type']
            sample_title = f"Sample {samp_name} {neg_cont_type} "
        
        elif samp_category == 'positive control':
            pos_cont_type = metadata_row['pos_cont_type']
            sample_title = f"Sample {samp_name} {pos_cont_type}"
        
        return sample_title